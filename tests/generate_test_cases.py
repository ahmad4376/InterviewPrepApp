#!/usr/bin/env python3
"""
Generate the supervisor's System Test Cases Excel deliverable.

Matches the exact template format from "Test Cases Template.xlsx":
  - Metadata rows (Test Case ID, Use Case Name, Priority, etc.)
  - Step table with merged B-C for User Actions
  - Test Case Execution Result row
  - Variation table: header row (T02-T05), pre/post conditions,
    then per-step Inputs + Expected Output rows with merged step-number cells
  - Variation Execution Result row

Output: tests/InterviewPrepApp_System_Test_Cases.xlsx

Run:
    python3 tests/generate_test_cases.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────
BOLD = Font(bold=True, size=11)
BOLD_12 = Font(bold=True, size=12)
NORMAL = Font(size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")

# Template column widths (A-I)
COL_WIDTHS = {
    "A": 22, "B": 17, "C": 32, "D": 34, "E": 35,
    "F": 34, "G": 32, "H": 16, "I": 23,
}


def _set_col_widths(ws):
    for letter, w in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w


def _cell(ws, row, col, value, font=NORMAL, alignment=WRAP, border=THIN):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = alignment
    c.border = border
    return c


def _border_range(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN


def _fill_range(ws, r1, c1, r2, c2, fill):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill


# ── Sheet builder ─────────────────────────────────────────────────────────────

def build_sheet(ws, uc):
    _set_col_widths(ws)
    row = 2

    # ── Note ──────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _cell(ws, row, 1,
          "Note: Write test cases for each use case separately. All of your "
          "test cases must be in a single excel file. However, test cases for "
          "each use case must be in a separate sheet in the file.",
          font=NORMAL, alignment=WRAP)
    row += 3  # skip to row 5

    # ── Metadata (rows 5-10) ──────────────────────────────────────────────
    # Row 5: Test Case ID | T01 | Use Case Name | <name>
    _cell(ws, row, 1, "Test Case ID", font=BOLD)
    _cell(ws, row, 2, "T01", font=BOLD)
    _cell(ws, row, 3, "Use Case Name", font=BOLD)
    _cell(ws, row, 4, uc["uc_name"], font=BOLD)
    _border_range(ws, row, 1, row, 9)
    row += 1

    # Row 6: Test Created by | name | Test executed by | <blank>
    _cell(ws, row, 1, "Test Created by", font=BOLD)
    _cell(ws, row, 2, uc["assigned"])
    _cell(ws, row, 3, "Test executed by")
    _cell(ws, row, 4, "")
    _border_range(ws, row, 1, row, 9)
    row += 1

    # Row 7: Test Case Priority | priority | Test Case Objectives | objectives
    _cell(ws, row, 1, "Test Case Priority", font=BOLD)
    _cell(ws, row, 2, uc["priority"])
    _cell(ws, row, 3, "Test Case Objectives", font=BOLD)
    _cell(ws, row, 4, uc["objectives"])
    _border_range(ws, row, 1, row, 9)
    row += 1

    # Row 8: Test browser/platform | value (merged B-D)
    _cell(ws, row, 1, "Test browser/platform", font=BOLD)
    _cell(ws, row, 2, "OS: Windows 10 / macOS / Linux, Browser: Google Chrome (latest)")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    _border_range(ws, row, 1, row, 9)
    row += 1

    # Row 9: Pre-conditions | value (merged B-D)
    _cell(ws, row, 1, "Pre-conditions", font=BOLD)
    _cell(ws, row, 2, uc["preconditions"])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    _border_range(ws, row, 1, row, 9)
    row += 1

    # Row 10: Post-conditions | value (merged B-D)
    _cell(ws, row, 1, "Post-conditions", font=BOLD)
    _cell(ws, row, 2, uc["postconditions"])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    _border_range(ws, row, 1, row, 9)
    row += 2  # blank row, then header

    # ── Step table header (row 12) ────────────────────────────────────────
    headers = ["Step No", "User actions", "", "Inputs", "System response",
               "Expected Outputs", "Actual Output", "Test Result (Pass/Fail)", "Comments"]
    for i, h in enumerate(headers, 1):
        _cell(ws, row, i, h, font=BOLD, alignment=CENTER_WRAP)
    # Merge B-C for "User actions"
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    _fill_range(ws, row, 1, row, 9, HEADER_FILL)
    _border_range(ws, row, 1, row, 9)
    header_row = row
    row += 1

    # ── Step rows ─────────────────────────────────────────────────────────
    steps = uc["steps"]
    for step in steps:
        step_no, action, inp, response, expected = step
        _cell(ws, row, 1, step_no, alignment=CENTER_WRAP)
        _cell(ws, row, 2, action)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        _cell(ws, row, 4, inp)
        _cell(ws, row, 5, response)
        _cell(ws, row, 6, expected)
        _cell(ws, row, 7, "")  # Actual Output — left blank for tester
        _cell(ws, row, 8, "")  # Test Result — left blank for tester
        _cell(ws, row, 9, "")  # Comments
        _border_range(ws, row, 1, row, 9)
        row += 1

    row += 1  # blank row

    # ── T01 Execution Result ──────────────────────────────────────────────
    _cell(ws, row, 1, "Test Case Execution Result", font=BOLD)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    _cell(ws, row, 3, "Passed")
    _cell(ws, row, 4, "Failed")
    _cell(ws, row, 5, "Not Executed")
    _border_range(ws, row, 1, row, 5)
    row += 2

    # ── Variations section ────────────────────────────────────────────────
    _cell(ws, row, 1,
          "Test Case Variations (Test cases with different inputs go here)",
          font=BOLD_12)
    row += 2

    variations = uc["variations"]  # list of dicts with keys: id, preconditions, postconditions, steps
    num_vars = len(variations)

    # Header row: | Test Case ID | T02 | T03 | T04 | T05 | ...
    _cell(ws, row, 2, "Test Case ID", font=BOLD, alignment=CENTER_WRAP)
    for vi, var in enumerate(variations):
        _cell(ws, row, 3 + vi, var["id"], font=BOLD, alignment=CENTER_WRAP)
    _border_range(ws, row, 2, row, 2 + num_vars)
    row += 1

    # Pre-conditions row
    _cell(ws, row, 2, "Pre-conditions", font=BOLD)
    for vi, var in enumerate(variations):
        _cell(ws, row, 3 + vi, var.get("preconditions", uc["preconditions"]))
    _border_range(ws, row, 2, row, 2 + num_vars)
    row += 1

    # Post-conditions row
    _cell(ws, row, 2, "Post-conditions", font=BOLD)
    for vi, var in enumerate(variations):
        _cell(ws, row, 3 + vi, var.get("postconditions", ""))
    _border_range(ws, row, 2, row, 2 + num_vars)
    row += 1

    # Step-number header
    _cell(ws, row, 1, "Step No", font=BOLD, alignment=CENTER_WRAP)
    _border_range(ws, row, 1, row, 1)
    row += 1

    # Per-step: Inputs row + Expected Output row, step no merged across both
    max_steps = max(len(var["steps"]) for var in variations)
    for si in range(max_steps):
        step_num = si + 1
        inp_row = row
        exp_row = row + 1

        # Merge step number across input + expected rows
        _cell(ws, inp_row, 1, step_num, alignment=CENTER_WRAP)
        ws.merge_cells(start_row=inp_row, start_column=1, end_row=exp_row, end_column=1)

        # Inputs label
        _cell(ws, inp_row, 2, "Inputs", font=BOLD)
        ws.merge_cells(start_row=inp_row, start_column=2, end_row=inp_row, end_column=2)

        # Expected Output label
        _cell(ws, exp_row, 2, "Expected Output", font=BOLD)
        ws.merge_cells(start_row=exp_row, start_column=2, end_row=exp_row, end_column=2)

        # Fill variation values
        for vi, var in enumerate(variations):
            col = 3 + vi
            if si < len(var["steps"]):
                s_inp, s_exp = var["steps"][si]
                _cell(ws, inp_row, col, s_inp)
                _cell(ws, exp_row, col, s_exp)
            else:
                _cell(ws, inp_row, col, "…")
                _cell(ws, exp_row, col, "…")

        _border_range(ws, inp_row, 1, exp_row, 2 + num_vars)
        row += 2  # next step pair

    # ── Variation Execution Result ────────────────────────────────────────
    _cell(ws, row, 1, "Test Case Execution Result", font=BOLD)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    for vi in range(num_vars):
        _cell(ws, row, 3 + vi, "Passed/Failed/Not executed")
    _border_range(ws, row, 1, row, 2 + num_vars)


# ── Use-case data ─────────────────────────────────────────────────────────────
# Each variation now has: id, preconditions, postconditions,
# steps = [(input, expected_output), ...]

USE_CASES = [
    {
        "sheet": "UC01 - Sign Up",
        "uc_name": "User Sign Up",
        "assigned": "Ahmad",
        "priority": "High",
        "objectives": "Check the functionality of user registration (sign up) use case.",
        "preconditions": "Application is accessible at base URL. User does not have an existing account. Clerk auth service is operational.",
        "postconditions": "User account created in Clerk. User authenticated and redirected to /dashboard.",
        "steps": [
            (1, "Navigate to Sign Up page", "URL: /sign-up", "Sign Up page loads with 'Get started' heading, Clerk form with email input, password input, and 'Continue with Google' button.", "Sign Up page displayed correctly."),
            (2, "Enter email address", "test@example.com", "Email input accepts the value. No validation errors.", "Email field populated."),
            (3, "Enter password", "StrongPassword123!", "Password input accepts value. Strength indicator shows acceptable.", "Password field populated."),
            (4, "Click 'Continue' button", "N/A", "Form submits. Clerk processes sign-up request.", "Account creation initiated."),
            (5, "Complete email verification", "Verification code from email", "Verification succeeds. Redirected to /dashboard.", "Dashboard page displayed with 'No interviews yet' empty state."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "User has a Google account. No existing app account.",
                "postconditions": "Account created via OAuth. Redirected to /dashboard.",
                "steps": [
                    ("Click 'Continue with Google'", "Google OAuth popup opens."),
                    ("Authenticate with Google", "Google authenticates user."),
                    ("N/A (auto redirect)", "Redirected to /dashboard. Dashboard displayed."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "User already has an account with email test@example.com.",
                "postconditions": "No new account created. User stays on sign-up page.",
                "steps": [
                    ("test@example.com", "Email input populated."),
                    ("StrongPassword123!", "Password input populated."),
                    ("Click 'Continue'", "Clerk shows 'email already in use' error message."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "User does not have an existing account.",
                "postconditions": "No account created. User stays on sign-up page.",
                "steps": [
                    ("newuser@example.com", "Email input populated."),
                    ("123 (weak password)", "Password input populated."),
                    ("Click 'Continue'", "Clerk shows password-strength validation error."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User does not have an existing account.",
                "postconditions": "No account created. User stays on sign-up page.",
                "steps": [
                    ("not-an-email (invalid format)", "Email input populated."),
                    ("Click 'Continue'", "Clerk shows 'invalid email address' validation error."),
                ],
            },
        ],
    },
    {
        "sheet": "UC02 - Sign In",
        "uc_name": "User Sign In",
        "assigned": "Ahmad",
        "priority": "High",
        "objectives": "Check the functionality of user authentication (sign in) use case.",
        "preconditions": "Application is accessible. User has a valid registered account. Clerk auth service is operational.",
        "postconditions": "User authenticated. Redirected to /dashboard. Session token stored.",
        "steps": [
            (1, "Navigate to Sign In page", "URL: /sign-in", "Sign In page loads with 'Welcome back' heading, Clerk form with identifier input, and 'Continue with Google' button.", "Sign In page displayed correctly."),
            (2, "Enter registered email", "user@example.com", "Identifier input populated. No errors.", "Email accepted."),
            (3, "Click 'Continue'", "N/A", "Password field appears (Clerk multi-step flow).", "Password step shown."),
            (4, "Enter correct password", "CorrectPassword123!", "Password input populated.", "Password accepted."),
            (5, "Click 'Continue' to sign in", "N/A", "Authentication succeeds. Redirected to /dashboard.", "Dashboard loads with 'Your Interviews' heading."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "User has account with email user@example.com.",
                "postconditions": "User stays on sign-in page. Error displayed.",
                "steps": [
                    ("user@example.com", "Identifier input populated."),
                    ("Click 'Continue'", "Password step shown."),
                    ("WrongPassword999!", "Password input populated."),
                    ("Click 'Continue'", "Clerk shows 'incorrect password' error."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "No account with the entered email exists.",
                "postconditions": "User stays on sign-in page. Error displayed.",
                "steps": [
                    ("nonexistent@example.com", "Identifier input populated."),
                    ("Click 'Continue'", "Clerk shows 'no account found' or 'couldn't find account' error."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "User is NOT authenticated (no session).",
                "postconditions": "User redirected to /sign-in. Protected page not accessible.",
                "steps": [
                    ("URL: /dashboard (direct navigation)", "Clerk middleware intercepts."),
                    ("N/A", "Redirected to /sign-in page automatically."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User has a Google account linked to the app.",
                "postconditions": "Authenticated via OAuth. Redirected to /dashboard.",
                "steps": [
                    ("Click 'Continue with Google'", "Google OAuth popup opens."),
                    ("Authenticate with Google", "Google authenticates user."),
                    ("N/A (auto redirect)", "Redirected to /dashboard."),
                ],
            },
        ],
    },
    {
        "sheet": "UC03 - Create Interview",
        "uc_name": "Create Interview",
        "assigned": "Abdul Samad",
        "priority": "High",
        "objectives": "Check the functionality of creating a new interview (technical or HR) use case.",
        "preconditions": "User is authenticated and on /create-interview. API server, MongoDB, and OpenAI API are operational.",
        "postconditions": "Interview created in DB with status 'scheduled'. Question pool generated. User redirected to /dashboard.",
        "steps": [
            (1, "Navigate to Create Interview page", "URL: /create-interview", "Page loads with 'Start from a Template' section, form fields (Job Title, Company, Job Description, Job Level, Number of Questions), Mass Interview toggle, and two submit buttons.", "Create Interview form displayed."),
            (2, "Enter job title", "Senior Backend Engineer", "Title input (#title) populated.", "Title field shows 'Senior Backend Engineer'."),
            (3, "Enter company name", "Acme Corp", "Company input (#company) populated.", "Company field shows 'Acme Corp'."),
            (4, "Enter job description", "We are looking for a Backend Engineer with Node.js and PostgreSQL expertise.", "Description textarea (#description) populated.", "Description populated."),
            (5, "Select job level", "Senior (from #jobLevel dropdown)", "Dropdown shows 'Senior'.", "Job level set to Senior."),
            (6, "Set number of questions", "5 (in #numQuestions)", "Input shows 5. Helper text: 'generate a larger pool (20 questions) and adaptively pick the best 5'.", "Questions set to 5."),
            (7, "Click 'Create Technical Interview'", "N/A", "Button shows 'Generating...' with spinner. After 10-15s, toast 'Interview created'. Redirect to /dashboard.", "Dashboard shows new interview with 'Scheduled' badge."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "User authenticated. On /create-interview page.",
                "postconditions": "HR interview created (type='hr'). Redirected to /dashboard.",
                "steps": [
                    ("Product Manager", "Title input populated."),
                    ("TechStart Inc", "Company input populated."),
                    ("Product Manager job description", "Description populated."),
                    ("Mid-Level", "Job Level dropdown set."),
                    ("5", "Number of questions set."),
                    ("Click 'Create HR Screening Interview'", "Button shows 'Generating...' then redirects to /dashboard."),
                    ("N/A", "Dashboard shows new interview with 'Scheduled' badge."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "User authenticated. On /create-interview page.",
                "postconditions": "Template auto-fills title and description. Ready to submit.",
                "steps": [
                    ("Click 'Frontend Developer' template card", "Card highlights green. 'Using template: Frontend Developer' shown."),
                    ("N/A (auto-filled)", "Title auto-fills to 'Frontend Developer'. Description auto-fills."),
                    ("WebCorp (company only)", "Company input populated. All fields now filled."),
                    ("Click 'Create Technical Interview'", "Both buttons now enabled. Submits and redirects to /dashboard."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "User authenticated. On /create-interview page. All fields empty.",
                "postconditions": "No interview created. User stays on create page.",
                "steps": [
                    ("Leave all fields empty", "Both 'Create HR Screening Interview' and 'Create Technical Interview' buttons are disabled (opacity-50, cursor-not-allowed)."),
                    ("Fill only title: 'Some Role'", "Buttons still disabled (company and description empty)."),
                    ("Fill company: 'SomeCorp'", "Buttons still disabled (description empty)."),
                    ("Fill description: 'A job description'", "Both buttons become enabled."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User authenticated. On /create-interview page.",
                "postconditions": "Number of questions clamped to valid range [1-20].",
                "steps": [
                    ("0 in #numQuestions", "Value clamps to 1 (min). Input shows 1."),
                    ("-5 in #numQuestions", "Value clamps to 1. Input shows 1."),
                    ("25 in #numQuestions", "Value clamps to 20 (max). Input shows 20."),
                    ("10 in #numQuestions", "Value accepted. Input shows 10."),
                ],
            },
        ],
    },
    {
        "sheet": "UC04 - Conduct Interview",
        "uc_name": "Conduct Voice Interview",
        "assigned": "Abdul Samad",
        "priority": "High",
        "objectives": "Check the functionality of conducting a voice interview (technical or HR) use case.",
        "preconditions": "User authenticated. A 'scheduled' interview exists. Microphone available. Deepgram API configured.",
        "postconditions": "Interview status → 'completed'. Transcript saved. Feedback generated at /feedback/[id].",
        "steps": [
            (1, "Navigate to Dashboard", "URL: /dashboard", "Dashboard loads. Scheduled interview shows 'Start' link button with arrow icon.", "'Start' button visible on scheduled interview card."),
            (2, "Click 'Start' on scheduled interview", "N/A", "Navigates to /interview/[id]. Pre-start screen shows interview title, company, '{N} questions prepared', and 'Start Interview' button.", "Interview pre-start screen displayed."),
            (3, "Click 'Start Interview'", "N/A", "Browser requests mic permission. Toast: 'Interview started — good luck!'. Active interview screen shows 'Question 1 of N' progress and 'End Interview' button.", "Active interview UI displayed."),
            (4, "Listen to AI interviewer's question", "N/A", "Deepgram voice agent speaks question via TTS (24kHz linear16 PCM). Audio plays through speaker.", "Question audio heard. Transcript updates with question text."),
            (5, "Answer the question verbally", "Spoken answer via microphone", "STT processes answer. AI evaluates and asks next question or concludes.", "Answer captured in transcript."),
            (6, "Click 'End Interview' after all questions", "N/A", "Confirmation: 'Are you sure? This will end the interview and generate feedback.' Click 'End & Generate Feedback'. Toast: 'Interview completed — generating feedback...'", "Interview ends. Redirected to feedback page."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Scheduled HR interview exists. Deepgram configured.",
                "postconditions": "HR interview completed. Feedback shows recommendation banner (Hire/Consider/Reject).",
                "steps": [
                    ("Click 'Start' on HR interview", "Navigates to /interview/[id]. Pre-start screen displayed."),
                    ("Click 'Start Interview'", "Active screen. AI asks HR-specific questions (communication, culture fit)."),
                    ("Answer HR questions verbally", "Answers captured. HR dimension scores generated."),
                    ("Click 'End Interview' → 'End & Generate Feedback'", "Feedback page shows recommendation banner and HR evaluation scores."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Scheduled interview exists. Microphone permission NOT granted.",
                "postconditions": "Interview cannot proceed. User stays on interview page.",
                "steps": [
                    ("Click 'Start' on interview", "Pre-start screen displayed."),
                    ("Click 'Start Interview'", "Browser requests mic permission."),
                    ("Click 'Deny' on mic prompt", "Microphone not available. App shows error or cannot record audio."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "An 'in-progress' interview exists on dashboard.",
                "postconditions": "Interview resumes from where it left off.",
                "steps": [
                    ("Click 'Resume' on in-progress interview", "Navigates to /interview/[id]. Interview session loads."),
                    ("N/A", "'Start Interview' or active interview screen shown."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User is on active interview page.",
                "postconditions": "Interview page reloads correctly.",
                "steps": [
                    ("Navigate to /dashboard (leave interview)", "Dashboard loads. Interview still in list."),
                    ("Navigate back to /interview/[id]", "Interview page reloads. 'Start Interview' or 'End Interview' button visible."),
                ],
            },
        ],
    },
    {
        "sheet": "UC05 - View Feedback",
        "uc_name": "View Interview Feedback",
        "assigned": "Abdul Rehman",
        "priority": "High",
        "objectives": "Check the functionality of viewing interview feedback use case.",
        "preconditions": "User authenticated. Completed interview with feedback exists in DB.",
        "postconditions": "Feedback displayed correctly. No data corruption.",
        "steps": [
            (1, "Navigate to Dashboard", "URL: /dashboard", "Dashboard loads. Completed interview shows 'View Feedback' link button.", "'View Feedback' link visible."),
            (2, "Click 'View Feedback'", "N/A", "Navigates to /feedback/[id]. Page shows interview title, company, date, 'Download PDF' link, and 'Dashboard' back link.", "Feedback page displayed."),
            (3, "Review 'Feedback' tab (default)", "N/A", "'Feedback' and 'Transcript' tab buttons visible. 'Feedback' tab active. 'Overall Score' section with score (e.g., 4.2/5).", "'Overall Score' displayed."),
            (4, "Review score breakdown", "N/A", "Component Scores section with bars for Correctness, Depth, Communication. Question-by-Question section with per-question scores.", "Score breakdown visible."),
            (5, "Review strengths and areas for improvement", "N/A", "'Strengths' section (green + markers). 'Areas for Improvement' section (amber ↑ markers).", "Both sections visible with content."),
            (6, "Click 'Transcript' tab", "N/A", "Tab switches to show full interview transcript with Q&A.", "Transcript displayed."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Completed HR interview with feedback exists.",
                "postconditions": "HR-specific feedback displayed.",
                "steps": [
                    ("Click 'View Feedback' on HR interview", "Feedback page loads."),
                    ("Review HR sections", "'Executive Summary', 'HR Evaluation Scores' sections visible."),
                    ("Review recommendation banner", "Banner shows Hire/Consider/Reject with description."),
                    ("Review HR scores", "Scores for Communication, Cultural Fit, Confidence, Clarity, Overall Suitability."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Completed interview with feedback. Feedback page loaded.",
                "postconditions": "PDF file accessible via link.",
                "steps": [
                    ("Click 'Download PDF' link", "Link opens /api/interviews/[id]/report in new tab (target='_blank')."),
                    ("N/A", "PDF report downloads or displays in browser."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "Completed interview but feedback NOT yet generated.",
                "postconditions": "User sees 'not available' message. Can refresh.",
                "steps": [
                    ("Navigate to /feedback/[id]", "Page shows 'Interview Feedback' heading."),
                    ("N/A", "'Feedback is not yet available for this interview.' and 'Feedback typically takes 10-15 seconds to generate.' displayed."),
                    ("Click 'Refresh' link", "Page reloads. If feedback now ready, shows scores."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User authenticated.",
                "postconditions": "Error or 404 displayed. No crash.",
                "steps": [
                    ("Navigate to /feedback/000000000000000000000000", "Server returns 404 or error."),
                    ("N/A", "'Feedback is not yet available' or error page shown."),
                ],
            },
        ],
    },
    {
        "sheet": "UC06 - Dashboard Filter",
        "uc_name": "Dashboard Search / Filter / Sort",
        "assigned": "Abdul Rehman",
        "priority": "Medium",
        "objectives": "Check the functionality of dashboard search, filter, and sort use case.",
        "preconditions": "User authenticated. Dashboard has multiple interviews with different statuses, titles, and companies.",
        "postconditions": "Filtered/sorted results correct. Original data unchanged. Clearing filters restores full list.",
        "steps": [
            (1, "Navigate to Dashboard", "URL: /dashboard", "Dashboard loads with 'Your Interviews' heading, search input (placeholder: 'Search by title or company...'), sort dropdown (Newest/Oldest/A-Z/Z-A), and status filter pills (All/Scheduled/In Progress/Completed).", "Dashboard toolbar displayed."),
            (2, "Type in search bar", "'Backend' in search field", "List filters in real-time. Only interviews with 'Backend' in title or company shown.", "Filtered results displayed."),
            (3, "Click 'Completed' status pill", "N/A", "List further filters to show only completed interviews matching search.", "Only completed interviews shown."),
            (4, "Select 'A-Z' from sort dropdown", "Sort: A-Z", "Visible interviews sorted alphabetically by title.", "Alphabetical order verified."),
            (5, "Clear search and click 'All' pill", "Clear search text, click 'All'", "Full interview list restored. Default sort applied.", "All interviews visible."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Dashboard has interviews. Search bar visible.",
                "postconditions": "Empty state shown. Clear filters restores list.",
                "steps": [
                    ("'xyznonexistent999' in search", "List filters in real-time."),
                    ("N/A", "'No interviews match your filters.' displayed."),
                    ("Click 'Clear filters'", "Search cleared. All interviews restored."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Dashboard has interviews with mixed statuses.",
                "postconditions": "Only completed interviews shown.",
                "steps": [
                    ("Click 'Completed' pill", "Filter applied."),
                    ("N/A", "Only cards with 'Completed' badge shown. Or 'No interviews match' if none completed."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "Dashboard has 2+ interviews created at different times.",
                "postconditions": "Sort order toggles correctly.",
                "steps": [
                    ("Select 'Oldest First'", "Interviews sorted by oldest creation date first."),
                    ("Select 'Newest First'", "Interviews sorted by newest creation date first. Order reverses."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "New user account with zero interviews.",
                "postconditions": "Empty state with CTA shown.",
                "steps": [
                    ("Navigate to /dashboard", "Dashboard loads."),
                    ("N/A", "'No interviews yet' heading and 'Create your first mock interview to get started.' text displayed."),
                    ("Click 'Create Interview' CTA button", "Navigates to /create-interview."),
                ],
            },
        ],
    },
    {
        "sheet": "UC07 - Edit Interview",
        "uc_name": "Edit Interview",
        "assigned": "Haider",
        "priority": "Medium",
        "objectives": "Check the functionality of editing an existing scheduled interview use case.",
        "preconditions": "User authenticated. Scheduled interview exists. Edit (pencil) button visible.",
        "postconditions": "Interview updated in DB. Dashboard reflects changes. Success toast shown.",
        "steps": [
            (1, "Navigate to Dashboard", "URL: /dashboard", "Dashboard loads. Scheduled interview card shows pencil icon button (aria-label: 'Edit interview').", "Edit button visible."),
            (2, "Click edit (pencil) button", "N/A", "API fetches interview details. Edit modal opens (role='dialog') with heading 'Edit Interview' and pre-filled inputs: #edit-title, #edit-company, #edit-description.", "Edit modal displayed with current values."),
            (3, "Modify the title", "Clear #edit-title, type 'Updated Interview Title'", "Title input updated.", "New title in input."),
            (4, "Modify the company", "Clear #edit-company, type 'New Corp'", "Company input updated.", "New company in input."),
            (5, "Click 'Save Changes'", "N/A", "PATCH request sent. 'Saving...' shown. Modal closes. Dashboard card shows 'Updated Interview Title' and 'New Corp'. Toast: 'Interview updated'.", "Updated values on dashboard."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Dashboard has interviews with mixed statuses.",
                "postconditions": "Edit button only on scheduled interviews.",
                "steps": [
                    ("Check scheduled interview card", "Pencil icon (aria-label: 'Edit interview') is visible."),
                    ("Check in-progress interview card", "Pencil icon is NOT visible."),
                    ("Check completed interview card", "Pencil icon is NOT visible."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Scheduled interview exists. Edit modal can be opened.",
                "postconditions": "No changes saved. Modal closed.",
                "steps": [
                    ("Click edit button → modal opens", "Modal displayed with current values."),
                    ("Change title to 'Should Not Be Saved'", "Title input modified."),
                    ("Click 'Cancel'", "Modal closes. Original title still shown on dashboard."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "Scheduled interview exists. Edit modal open.",
                "postconditions": "Save rejected. Modal stays open or error shown.",
                "steps": [
                    ("Clear #edit-title, type '   ' (spaces)", "Title input contains whitespace."),
                    ("Click 'Save Changes'", "Server rejects. Error displayed or modal remains open."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User authenticated. Knows another user's interview ID.",
                "postconditions": "Request rejected. No data modified.",
                "steps": [
                    ("Send PATCH /api/interviews/[other-id] with modified data", "API returns 403 Forbidden or 404 Not Found."),
                    ("N/A", "No changes made to the other user's interview."),
                ],
            },
        ],
    },
    {
        "sheet": "UC08 - Delete Interview",
        "uc_name": "Delete Interview",
        "assigned": "Haider",
        "priority": "Medium",
        "objectives": "Check the functionality of deleting an interview use case.",
        "preconditions": "User authenticated. At least one interview exists. Delete (trash) button visible.",
        "postconditions": "Interview removed from DB. Card removed from dashboard. Toast: 'Interview deleted'.",
        "steps": [
            (1, "Navigate to Dashboard", "URL: /dashboard", "Dashboard loads. Each interview card has trash icon button (aria-label: 'Delete interview').", "Delete buttons visible on all cards."),
            (2, "Click delete (trash) button", "N/A", "Confirmation banner appears below card: 'Delete this interview? This cannot be undone.' with 'Cancel' and 'Delete' buttons.", "Confirmation dialog visible."),
            (3, "Click 'Delete' to confirm", "N/A", "DELETE request sent. Button shows 'Deleting...'. Card removed from list. Toast: 'Interview deleted'.", "Interview removed from list."),
            (4, "Verify count decreased", "N/A", "Dashboard list has one fewer card. If last interview, 'No interviews yet' empty state appears.", "List updated correctly."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Interview exists. Delete confirmation shown.",
                "postconditions": "Interview NOT deleted. Confirmation hidden.",
                "steps": [
                    ("Click trash icon → confirmation appears", "'Delete this interview? This cannot be undone.' shown."),
                    ("Click 'Cancel'", "Confirmation hides. Interview card still present."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Dashboard has interviews of all three statuses.",
                "postconditions": "Delete button present on all cards.",
                "steps": [
                    ("Check scheduled interview card", "Trash icon (aria-label: 'Delete interview') visible."),
                    ("Check in-progress interview card", "Trash icon visible."),
                    ("Check completed interview card", "Trash icon visible."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "Interview exists. Title noted before deletion.",
                "postconditions": "Interview title no longer in DOM.",
                "steps": [
                    ("Note first interview title", "Title recorded (e.g., 'Backend Interview')."),
                    ("Click trash → 'Delete'", "Card removed. Toast shown."),
                    ("Search for deleted title", "Title no longer visible on dashboard."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "User authenticated. Knows another user's interview ID.",
                "postconditions": "Request rejected. No data deleted.",
                "steps": [
                    ("Send DELETE /api/interviews/[other-id]", "API returns 403 Forbidden or 404 Not Found."),
                    ("N/A", "No interview deleted from other user's account."),
                ],
            },
        ],
    },
    {
        "sheet": "UC09 - Coding Interview",
        "uc_name": "Coding Interview",
        "assigned": "Kabli",
        "priority": "High",
        "objectives": "Check the functionality of the coding interview (LeetCode-style) use case.",
        "preconditions": "User authenticated. Coding problems in DB (LeetcodeQuestion). Code execution API (Piston) operational. Monaco Editor loads.",
        "postconditions": "Code executed. Test results displayed (pass/fail). Expected vs actual output shown.",
        "steps": [
            (1, "Navigate to Coding Interview", "URL: /coding-interview", "Page loads with problem title (h1), difficulty badge, problem counter ('1 / N'), language selector, 'Description' tab active, Monaco code editor, and 'Run Code' button.", "Coding interview UI displayed."),
            (2, "Read problem statement", "N/A", "'Description' tab shows problem description, examples (Input/Output), time limit, memory limit, and tags.", "Problem statement visible."),
            (3, "Write correct solution in editor", "Type or paste code in Monaco editor", "Editor accepts input. Syntax highlighting active.", "Code entered in editor."),
            (4, "Click 'Run Code'", "N/A", "Button changes to 'Running...' with spinner. Code sent to Piston execution API.", "Execution in progress."),
            (5, "Review test results", "N/A", "'Test Results' panel appears. Shows 'X/Y passed'. Each case shows 'Accepted' (green) badge.", "All test cases passed."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Coding page loaded. Language selector visible.",
                "postconditions": "Language switched. Editor updates syntax highlighting.",
                "steps": [
                    ("Select 'Python' from language dropdown", "Editor switches to Python. Syntax highlighting updates."),
                    ("Select 'C++' from language dropdown", "Editor switches to C++."),
                    ("Select 'JavaScript' from language dropdown", "Editor switches back to JavaScript."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Coding page loaded. Problem displayed.",
                "postconditions": "Failed test cases shown with expected vs actual.",
                "steps": [
                    ("Write incorrect code: 'return -1'", "Code entered in editor."),
                    ("Click 'Run Code'", "Button shows 'Running...'. Results returned."),
                    ("N/A", "'Wrong Answer' badge shown. 'Expected:' and 'Your output:' labels show diff."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "Coding page loaded. Problem displayed.",
                "postconditions": "Error output shown in results.",
                "steps": [
                    ("Write code with syntax error: 'function() {{{' ", "Code with errors entered."),
                    ("Click 'Run Code'", "Button shows 'Running...'. Results returned."),
                    ("N/A", "'Error:' label shown in test results with error message."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "Coding page loaded. Multiple problems exist (N > 1).",
                "postconditions": "Problem counter and content update correctly.",
                "steps": [
                    ("Verify '← Prev' NOT visible on first problem", "Counter shows '1 / N'. '← Prev' hidden."),
                    ("Click 'Next →'", "Counter updates to '2 / N'. Problem description changes. '← Prev' appears."),
                    ("Click '← Prev'", "Counter back to '1 / N'. Original problem restored."),
                ],
            },
        ],
    },
    {
        "sheet": "UC10 - Mass Interview",
        "uc_name": "Mass Interview Flow",
        "assigned": "Kabli",
        "priority": "Medium",
        "objectives": "Check the functionality of the mass interview (shared link) flow use case.",
        "preconditions": "Creator user authenticated. Mass Interview toggle on create form. MongoDB, Deepgram, OpenAI operational.",
        "postconditions": "Mass interview created with shareToken. Candidate sessions tracked. Creator views results.",
        "steps": [
            (1, "Navigate to Create Interview", "URL: /create-interview", "Form loads with Mass Interview toggle switch (role='switch').", "Toggle visible."),
            (2, "Fill form and enable Mass Interview", "Title: 'Mass QA Interview', Company: 'TestCorp', Description: '...', Toggle: ON", "Toggle turns green (aria-checked='true'). Sub-label: 'Allow multiple candidates to take this interview via a shareable link'.", "Form ready."),
            (3, "Click 'Create Technical Interview'", "N/A", "'Generating...' then toast 'Interview created'. Redirected to /dashboard.", "Dashboard shows new mass interview."),
            (4, "Verify mass-interview UI on dashboard", "N/A", "Card shows purple 'Mass' badge, 'Copy Link' button, and 'Candidates' link. No 'Start' button.", "Mass interview UI visible."),
            (5, "Click 'Copy Link'", "N/A", "Share URL copied to clipboard. Button changes to 'Copied!' with green check. Toast: 'Invite link copied to clipboard'.", "Link copied confirmation."),
            (6, "Candidate navigates to /join/[token]", "Share link URL", "Join page loads with interview title, company, description, question count, and 'Start Interview' button.", "Candidate join page displayed."),
            (7, "Creator clicks 'Candidates' link", "N/A", "Navigates to /interviews/[id]/candidates. Lists candidate sessions with status.", "Candidates page displayed."),
        ],
        "variations": [
            {
                "id": "T02",
                "preconditions": "Creator has a mass interview. Creator is still authenticated.",
                "postconditions": "Creator blocked from joining own interview.",
                "steps": [
                    ("Creator navigates to /join/[own-token]", "Join page loads or error shown."),
                    ("N/A", "'Unable to Join' heading or error: creator cannot be a candidate."),
                ],
            },
            {
                "id": "T03",
                "preconditions": "Invalid or non-existent share token.",
                "postconditions": "Error page displayed. No session created.",
                "steps": [
                    ("Navigate to /join/invalid-token-12345", "Page loads."),
                    ("N/A", "'Unable to Join' heading with error message (invalid or not found)."),
                ],
            },
            {
                "id": "T04",
                "preconditions": "Candidate previously completed the interview.",
                "postconditions": "Completed state shown. Can view feedback.",
                "steps": [
                    ("Candidate revisits /join/[token]", "Join page loads."),
                    ("N/A", "'Already Completed' heading with 'You have already completed this interview.' text."),
                    ("Click 'View Feedback'", "Navigates to candidate feedback page."),
                ],
            },
            {
                "id": "T05",
                "preconditions": "Candidate left mid-interview. Session is 'in-progress'.",
                "postconditions": "Session resumes from where left off.",
                "steps": [
                    ("Candidate revisits /join/[token]", "Join page loads."),
                    ("N/A", "'Interview In Progress' heading with 'Resume where you left off.' text."),
                    ("Click 'Resume Interview'", "Navigates to interview session. Interview continues."),
                ],
            },
        ],
    },
]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    for uc in USE_CASES:
        ws = wb.create_sheet(title=uc["sheet"])
        build_sheet(ws, uc)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "InterviewPrepApp_System_Test_Cases.xlsx")
    wb.save(out_path)
    print(f"Excel file generated: {out_path}")


if __name__ == "__main__":
    main()
